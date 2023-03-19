import os
import threading
import time
import datetime
import telebot
from fuzzywuzzy import fuzz
from openpyxl import Workbook, load_workbook

from Controllers.BotTg import BotTg
from Controllers.actions import execute_cmd
from Controllers.threads import thread_conditioner, emulator_degrees, protection_func
from constant import GET_TEMP, SET_TEMP, TURN_ON_LIGHT, TURN_OFF_LIGHT, OPTS
from variables import lights, currentTemp, conditionerState, botObj, user_id, excel


wb = Workbook()

sheet = wb.active

sheet.append(['Температура','Кімната','Вітальня','Спальня','Кухня','Вбиральня','Зала','Кондиціонер','Статус'])

wb.save("res.xlsx")

wb = load_workbook('res.xlsx')
global excel
excel['wb'] = wb

time.sleep(1)

x = threading.Thread(target=emulator_degrees, args=())
x.start()

x = threading.Thread(target=protection_func, args=())
x.start()

global botObj
botObj['bot'] = telebot.TeleBot('6004805138:AAF4lpUfT63_coSSh9gy1254ghQxHOrB89c')
print('aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa', botObj)
b = BotTg(botObj['bot'])

while True:
    try:
        botObj['bot'].polling(none_stop=True, interval=0)
    except Exception as _ex:
        print(_ex)
        time.sleep(15)

# callback('Іван в якому стані зараз кондиціонер?')
# callback('Іван яка зараз температура?')
# callback('Іван які лампи зараз увімкнено')
# callback('Іван увімкни світло у Кухні')
# callback('Іван увімкни світло у Кімнаті')
# callback('Іван які лампи зараз увімкнено')
# callback('Іван зроби температуру 31')
#
# time.sleep(9)
# callback('Іван яка зараз температура?')
# callback('Іван в якому стані зараз кондиціонер?')
# callback('Іван увімкни світло у всій квартирі')
# callback('Іван почалась комендантська година')
# time.sleep(9)
# callback('Іван яка зараз температура?')
# callback('Іван в якому стані зараз кондиціонер?')
# time.sleep(9)
# callback('Іван яка зараз температура?')
# callback('Іван в якому стані зараз кондиціонер?')
# callback('Іван в якому стані зараз квартира?')
#
# time.sleep(20)
